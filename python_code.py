import PyPDF4
import os
import openpyxl
import re

# the sorted_alphanumeric function is taken from https://stackoverflow.com/questions/4836710/is-there-a-built-in-function-for-string-natural-sort
def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [ convert(c) for c in re.split('([0-9]+)', key) ]
    return sorted(data, key=alphanum_key)

def list_pdf_page(my_directory, type_file, output_name):
    my_path = Path(my_directory)
    list_files = sorted_alphanumeric(os.listdir(my_path))
    my_len_list = len(list_files)
    list_interest = []
    file_interest_list = []
    for i in range(0,my_len_list):
        file_path = my_path/list_files[i]
        file_suffix = file_path.suffix
        if file_suffix == type_file:
            try:
                my_pdfFileObj = open(file_path, 'rb')
                my_pdfReader = PyPDF4.PdfFileReader(my_pdfFileObj)
                pages = my_pdfReader.getNumPages()
                list_interest = list_interest + [pages]
                file_interest_list = file_interest_list + [file_path.stem]
            except:
                print("Error reading '{}'".format(file_path))
                continue
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    length_list = len(list_interest)
    sheet.cell(row=1, column=1).value = 'S/N'
    sheet.cell(row=1, column=2).value = 'File Name'
    sheet.cell(row=1, column=3).value = 'No. of Pages'
    for i in range(1, length_list + 1):
        sheet.cell(row=i+1, column=1).value = i
        sheet.cell(row=i+1, column=2).value = file_interest_list[i-1]
        sheet.cell(row=i+1, column=3).value = list_interest[i-1]
    wb.save(output_name)
